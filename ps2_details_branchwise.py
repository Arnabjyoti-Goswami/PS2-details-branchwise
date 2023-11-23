import pandas as pd
import os
from openpyxl.styles import Alignment, Font, Border, Side

nb_dir = os.path.abspath('')
pd.set_option('display.max_rows', 500)

sort_by_col = 'Stipend'
branches_col = 'Preferred Branches'

# Show details regarding a df for debugging (not used in here, was used in the jupyter notebook which was used to develop this script)
def show_df_details(df, df_name, num_rows=5):
  print(f"Number of rows in the dataframe '{df_name}' = {len(df)}")
  display(df.head(num_rows))


# Convert strings to numeric datatypes
def convert_to_numeric(entry):
  # Check if the entry contains digits or a single decimal point
  if entry.replace('.', '', 1).isdigit():
    # Convert to integer if only digits are present
    if '.' not in entry:
      return int(entry)
    # Convert to float if a single decimal point is present
    elif entry.count('.') == 1:
      return float(entry)

  # Return the entry without any change, if its a non-numeric datatype
  return entry


def sort_df_column(df, sort_column):
  '''
  Sort a dataframe based on a single column's values, non numeric values are put at the bottom
  sorting is done in descending order, can make ascending=True to make it ascending order
  '''
  if sort_column in df.columns:
    # Create a mask for non-numeric values
    non_numeric_mask = ~df[sort_column].apply(lambda x: isinstance(x, (int, float)))

    # Sort the DataFrame using the mask
    df_numeric = df[~non_numeric_mask].sort_values(by=sort_column, ascending=False)
    df_non_numeric = df[non_numeric_mask]

    # Concatenate the sorted DataFrames to combine numeric and non-numeric values
    sorted_df = pd.concat([df_numeric, df_non_numeric])

    # Reset the index to maintain a continuous index
    sorted_df.reset_index(drop=True, inplace=True)

    return sorted_df
  else:
    print(f"Warning: Column '{sort_column}' not found in the DataFrame. Skipping sorting...")
    return df
  

def filter_single_branch_df(df, branch):
  '''
  Returns a dataframe that is filtered according to the required branch (only for rows in the dataframe where the 'Preferred Branches'
  column contains the required branch, if there are no such rows then a filtered dataframe cannot be made, so return None instead).
  The sorting and showing details functions are also called within this function.
  '''
  try:
    # filter to make a new dataframe where the rows CONTAINS the branch required
    # Note: filter = (df[branches_col] == filter_single_branch)    - if you want to find the rows for which
    # there is ONLY the branch required and not CONTAINS the branch required with other branches
    filter = (
        ( df[branches_col].str.contains(branch) )
        &
        ( df[branches_col].str.count(branch) == 1 )
        &
        # for the branch 'Any', otherwise it will consider 'AnyA7' type branches as also 'Any'
        ( df[branches_col].str.contains(branch + ' ,') | df[branches_col].str[-len(branch):].str.contains(branch) )
        &
        # remove the rows where there are there are dual degrees along with the branch,
        # as the dual degrees would be done separately. If needs to be done together, then
        # remove this condition and later remove the dual_degrees from the all_degrees (in the last cell)
        ( ~df[branches_col].str.contains('Any' + branch) )
    )

    filtered_df = df[filter]

    filtered_df = sort_df_column(filtered_df, sort_by_col) # Sorting the filtered_df here itself

    # show_df_details(filtered_df, branch) # Showing Details of the filtered_df here itself

    return filtered_df

  except:
    print(f"No row in the dataframe 'df' has {branches_col} as {branch}")
    return
  

# Generate and save an excel file that has different sheets for different branches
def save_excel(df, branches, output_excel_filename='Output.xlsx'):
  output_excel_filepath = os.path.join(nb_dir, output_excel_filename)

  branch_wise_filtered_df_s = []
  for branch in branches:
    filtered_df_for_single_branch = filter_single_branch_df(df, branch)
    if filtered_df_for_single_branch is not None:
      branch_wise_filtered_df_s.append(filtered_df_for_single_branch)

  # Save the initial excel file without any formatting
  writer = pd.ExcelWriter(output_excel_filepath, engine='openpyxl')
  for i, df in enumerate(branch_wise_filtered_df_s):
    df.to_excel(writer, sheet_name = branches[i], index=False)

  # Then load the excel file with the data for formatting
  workbook = writer.book

  for i, df_branch in enumerate( branch_wise_filtered_df_s ):
    worksheet_name = branches[i]
    worksheet = workbook[worksheet_name]

    # Adjust the column width for all columns
    for column in worksheet.columns:
      max_length = 0
      column_name = column[0].column_letter  # Get the column name (e.g., 'A', 'B', 'C', ...)

      for cell in column: # Find the length of the longest content in each column
        if len(str(cell.value)) > max_length:
          max_length = len(cell.value)

      adjusted_width = (max_length + 1) # adjusts the width of all columns in the Excel worksheet to accommodate the longest content in each column. You can adjust the adjusted_width calculation to fit your specific formatting needs.
      worksheet.column_dimensions[column_name].width = adjusted_width

    # Adjust the row height and align all cells to center for all rows from 2nd to last
    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
      for cell in row:
        cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
        worksheet.row_dimensions[cell.row].height = 20

    # Style the 1st row with a bigger height and bold text
    for row in worksheet.iter_rows(min_row=1, max_row=1):
      for cell in row:
        cell.font = Font(bold=True)  # Set text to bold
        cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center') # Center it
        worksheet.row_dimensions[cell.row].height = 30  # Set a bigger row height

    # Create and apply a default black line border to all cells
    border = Border(
      left=Side(style='thin', color='000000'),
      right=Side(style='thin', color='000000'),
      top=Side(style='thin', color='000000'),
      bottom=Side(style='thin', color='000000')
    )
    for row in worksheet.iter_rows():
      for cell in row:
        cell.border = border
  workbook.save(output_excel_filename)
  print('\n')
  print(f"{output_excel_filename} saved to '{output_excel_filepath}' successfully")


if __name__ == "__main__":
  # Read in the csv input file:
  file_name = 'StationDetails.csv' # Input file name
  file_path = os.path.join(nb_dir, file_name)
  df = pd.read_csv(file_path)

  df.rename(
    columns={
      'Stipend (UG)': sort_by_col,
      'Preferred Branches': branches_col,
    },
    inplace=True,
  )

  df[branches_col] = df[branches_col].str.strip()
  df[sort_by_col] = df[sort_by_col].apply(convert_to_numeric)
  df = sort_df_column(df = df, sort_column = sort_by_col)
  df['Stipend (PG)'] = df['Stipend (PG)'].apply(convert_to_numeric)
  # show_df_details(df, 'df')

  # Make excel output file:
  single_degrees = ['A1', 'A2', 'A3', 'A4', 'A5', 'A7', 'A8', 'AA', 'AB', 'B1', 'B2', 'B3', 'B4', 'B5']
  dual_degrees = ['Any' + branch for branch in single_degrees]
  all_degrees = single_degrees + dual_degrees
  all_degrees.insert(0, 'Any')
  all_degrees.append('Unavailable')

  excel_output_filename = 'Branchwise PS2 Station Details.xlsx'

  save_excel(
      df = df,
      branches = all_degrees,
      output_excel_filename = excel_output_filename,
  )

